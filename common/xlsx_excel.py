# coding=UTF-8

import openpyxl
import xlrd
from xlrd import sheet
import pandas as pd

from conf.all_path import hstq_file_path_dxk, mpcr_result_consistence_amt_import_path


def write_excel_xlsx(path, value1, value2, colname1, colname2):
    """
    把各流程生成的数据写入Excel，需要两个参数，实验流程模块传入lims号和实验室号，订单模块传订单号和患者名称
    :param colname2: 列名2
    :param colname1: 列名1
    :param path: 写入文件的路径
    :param value1:传入find_elements方法获取的多个元素，直接传入即可
    :param value2:find_elements方法获取的多个元素，直接传入eles即可(eles=find_elements(ele))
    :return:
    """
    # 二维list
    company_name_list = []
    value_dict = dict(zip(value1, value2))
    for i, j in value_dict.items():
        # print(i.text, j.text)
        company_name_list.append([i.text, j.text])

    df = pd.DataFrame(company_name_list, columns=[colname1, colname2])
    # print(company_name_list)
    # 保存到本地excel
    df.to_excel(path, index=False)
    print("表格写入数据成功！")


# excel中获取待选样本
def get_lims_for_excel_by_col(path):
    """
    从对应的Excel中获取上一步流传下来的本节点的待选表lims样本号,整列获取
    """
    sample_nub = read_excel_col(path, 'lims号')
    lims_id_str = "\n".join(sample_nub)
    print(lims_id_str)
    return lims_id_str


def get_lims_for_excel_by_rows(path, s_row, e_row, colName):
    """
    从对应的Excel中获取上一步流传下来的本节点的待选表lims样本号,整列获取
    """
    sample_nub = read_excel_col_by_row(path, s_row, e_row, colName)
    lims_id_str = "\n".join(sample_nub)
    # print(lims_id_str)
    return lims_id_str


def read_excel_col(path, colName):
    """
    根据列名，读取指定Excel的列，输出为list
    :param path: excel文件名
    :param colName: 要读取的列名
    :return: 返回列组成的list，['','','','']
    """
    df = pd.read_excel(path, keep_default_na=False)  # 使用pd读取excel
    df_data = df[colName].values.tolist()
    # print(df_data)
    return df_data


def read_excel_col_by_row(path, s_row, e_row, colName):
    """
    读取指定行，指定列的数据
    :param s_row: 行切片开始下标
    :param e_row: 行切片结束下标
    :param path: excel文件名
    :param colName: 列名
    :return: 返回列组成的list，['','','','']
    """
    df = pd.read_excel(path, keep_default_na=False)  # 使用pd读取excel
    data1 = df.loc[df.index[s_row:e_row], colName].values.tolist()
    return data1


def read_excel_xlsx_list_col(path, sheet_index, colname):
    """
    获取指定sheet页&指定列的数据，返回list
    :param sheet_index: 页索引位置，从0开始
    :param path:文件路径
    :param colname:指定列名，不是列数

    """
    row_list, col_list, ret = [], [], []
    data = xlrd.open_workbook(path)  # 获取文本对象
    table = data.sheets()[sheet_index]  # 根据index获取某个sheet
    rows = table.nrows  # 获取当前sheet页面的总行数,把每一行数据作为list放到 list
    df = pd.read_excel(path)  # 使用pd读取excel
    col_name = list(df)  # 获取列名
    while colname in col_name:
        col = col_name.index(str(colname))
        # print("表格中，" + str(colname) + "，处于第" + str(col_name.index(str(colname)) + 1) + "列")
        # 获取指定列的每行信息（除去列名那行不算）
        for i in range(1, rows):
            ret.append([table.cell_value(i, col)])
        break
    return ret


def write_excel_xlsx_by_openpyxl(path, sheet_name, value):
    """
    新建实验流程开始时的样本流转记录表，记录从样本处理开始，各实验流程结果表中样本的下一步流程，openpyxl操作单元格行列下标从（1,1）开始
    :param path: 建立对应下一步流程的Excel表
    :param sheet_name: 记录表sheet页名
    :param value: 要记录的样本数据，以列表嵌套列表格式，子列表数量不限，[[''],[''],[''],...]
    """
    index = len(value)
    workbook = openpyxl.Workbook()  # 新建工作簿（默认有一个sheet）
    sheet_num = workbook.active  # 获得当前活跃的工作页，默认为第一个工作页
    sheet.title = sheet_name + '明细表数据'  # 给sheet页的title赋值
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet_num.cell(row=i + 1, column=j + 1, value=str(value[i][j]))  # 行，列，值 这里是从1开始计数的
    workbook.save(path)  # 一定要保存
    print("初始化数据存放Excel文件成功！")


def pandas_write_excel(company_name_list, path):
    """
    pandas写入Excel
    :param company_name_list:二维列表
    :param path: 文件保存位置
    """
    # list转dataframe
    df = pd.DataFrame(company_name_list)

    # 保存到本地excel，不带表头
    df.to_excel(path, header=None, index=False)


def add_write_excel_xlsx(path, value):
    """
    往各实验流程样本流转记录表中追加数据，对在不同实验流程节点结果表生成的不同的下一步样本，追加到对应的流程记录表中
    :param path: 对应流程记录表
    :param value: 要记录的样本数据，以列表嵌套列表格式，子列表数量不限，[[''],[''],[''],...]
    """
    data = openpyxl.load_workbook(path)
    # 取第一张表

    table = data.active
    print(table.title)  # 输出表名
    nrows = table.max_row  # 获得行数
    # 注意行业列下标是从1开始的
    for i in range(1, len(value) + 1):
        for j in range(1, len(value[i - 1]) + 1):
            table.cell(nrows + i, j).value = str(value[i - 1][j - 1])
            # print('写入：', value[i - 1][j - 1])
    data.save(path)


def getDataByRowColumn(path, rows, column):
    """
    获取指定单元格的值，行列下标去除标题，从【0,0】开始
    :param path: 文件
    :param rows: 行号
    :param column: 列号
    :return: 单元格值
    """
    data = pd.read_excel(path, header=0)
    ordreNumb = data.iloc[rows, column]
    return str(ordreNumb)

def openpyxl_edit_data(filepath,datas,nub):
    """openpyxl打开并清空excel，之后在指定位置写入数据"""

    # 打开Excel文件
    workbook = openpyxl.load_workbook(filepath)
    # 选择第一个工作表
    worksheet = workbook.worksheets[0]
    # 清空所有单元格的内容
    for row in worksheet:
        for cell in row:
            cell.value = None
    # 在B2单元格写入数据
    for i in range(len(datas)):
        worksheet.cell(i+1, 1, datas[i])  # 修改第i+1行，第index列值
        worksheet.cell(i+1, 2, nub+i)  # 修改第i+1行，第index列值
    # 保存修改后的Excel文件
    workbook.save(filepath)




if __name__ == '__main__':
    lists=[['GSD230511024B001', '1'], ['GSD230519026B001', '2']]


