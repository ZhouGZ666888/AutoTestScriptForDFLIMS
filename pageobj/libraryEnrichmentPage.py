# -*- coding: utf-8 -*-
# @Time    : 2022/08/28
# @Author  : guanzhong.zhou
# @File    : 文库富集模块页面功能封装
from datetime import datetime
import pyperclip
import xlrd
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
from common.screenshot import Screenshot
from common.DataBaseConnection import executeSql
from conf.all_path import wkfj_file_path, functionpageURL_path, wkfj_detail_import_path, wkfj_result_import_path, \
    position_in_box_path, wkfj_probe_file_path
from conf.config import pooling_result
from common.editYaml import *
from common.xlsx_excel import read_excel_col, get_lims_for_excel_by_col, pandas_write_excel, write_data_toexcle
from PageElements.wkfj_ele import *
from conf.execute_sql_action import wkfj_detail_sql1, fj_next_step, wkfj_detail_sql2, wkfj_detail_sql3
from uitestframework.basepageTools import BasePage
from common.logs import log


def set_preinstall_throughput():
    """设置预期通量"""
    order = get_order()
    executeSql.test_updateByParam(wkfj_detail_sql2.format(order))


class LibraryenrichmentPage(BasePage):
    """文库富集页面方法封装"""

    # 获取页面提示信息
    def get_pageinfo(self):
        """
        获取页面操作提示信息
        Task list saved successfully---保存样本到任务单成功
        Submit successfully---提交成功
        sample in storage successfully---入库成功
        """
        return self.get_text('xpath', page_success_info)

    # 新增任务单
    def add_task(self, taskType):
        """新建文库富集任务单"""
        if taskType == '无探针Pooling':
            set_preinstall_throughput()

        log.info("文库富集首页，点击新建按钮，进入样本待选表，新增{}任务".format(taskType))
        self.clicks('css', add_sample_process_task)
        self.wait_loading()
        self.wait_loading()

        log.info("选择任务类型")
        self.clicks('css', task_type)
        self.sleep(0.5)
        self.clicks('xpath', task_type_choice.format(taskType))
        self.wait_loading()

        log.info("录入任务描述")
        self.input('css', task_description, '文库富集测试')
        self.sleep(0.5)

        log.info("选择操作方式")
        self.clicks('css', action_stlye)
        self.clicks('css', action_stlye_choice)
        self.sleep(0.5)
        log.info("选择sop")
        self.clicks('css', select_sop)
        self.sleep(0.5)
        self.clicks('css', select_sop_choice)
        self.wait_loading()

    # 待选表校验lims号
    def get_non_probe_sample(self):
        """无探针Pooling任务类型样本数据"""
        lims_id_str = get_lims_for_excel_by_col(wkfj_file_path)
        log.info('获取接样流程中的lims样本号,核对lims号是否存在')
        self.clicks('css', check_lims_sample_num)
        self.sleep(1)
        return self.search_and_add_sample(lims_id_str)

    def get_probe_sample(self):
        """有探针任务类型样本数据"""
        # 将有探针任务类型的样本存到Excel中
        order = get_order()
        ata = executeSql.test_select_limsdb(wkfj_detail_sql3.format(order))
        values_list = [list(d.values()) for d in ata]
        print(values_list)
        write_data_toexcle(wkfj_probe_file_path, values_list)

        lims_id_str = get_lims_for_excel_by_col(wkfj_probe_file_path)
        log.info('获取接样流程中的lims样本号,核对lims号是否存在')
        self.clicks('css', check_lims_sample_num)
        self.sleep(1)
        return self.search_and_add_sample(lims_id_str)

    def search_and_add_sample(self, sampledata):
        """
        待选表核对lims号功能，并保存任务单号
        :param sampledata: 样本信息
        :return: 添加样本成功信息
        """
        self.input('css', check_lims_sample_number_textarea, sampledata)
        self.sleep(0.5)
        self.clicks('css', check_lims_sample_number_confirm)
        self.wait_loading()
        if self.isElementExists('xpath',result_check):
            self.click_by_js('xpath',result_check)
            self.sleep(0.5)
        log.info('选中核对后的样本，点击【加入选中样本&保存】')
        self.click_by_js('css', all_choice)
        self.sleep(1)
        self.click_by_js('css', addSelect_or_save_btn)
        pageinfo = self.get_pageinfo()
        Screenshot(self.driver).get_img("文库富集待选表点击核对lims号，录入样本号进行查询，勾选查询结果，并保存任务单号", "保存任务单成功")
        self.wait_loading()
        return pageinfo

    # 进入明细表或结果表
    def enter_result_list(self, ele, page):
        """
        待选表或者明细表，在进入下一节点时，获取下一节点URL地址，并写入临时数据文件中
        :param ele: 点击进入下一节点元素定位
        :param page: 下一节点页面名称
        """
        urldata = read_yaml(functionpageURL_path)
        self.sleep(1)
        self.clicks('css', ele)
        log.info('点击按钮进入{}'.format(page))
        self.wait_loading()
        self.sleep(2)
        url = self.get_current_url()  # 获取当前页面URL地址
        urldata["url"] = url
        save_yaml(functionpageURL_path, urldata)
        print("写入后的URL地址", urldata)

    # 准备批量粘贴导入的数据
    def read_import_data(self, probe):
        """读取Excel中批量粘贴导入的数据"""
        str_time = datetime.now().strftime('%Y%m%d')  # 获取当前时间

        samples = self.findelements('css', all_samples)
        wb = load_workbook(filename=wkfj_detail_import_path)  # 打开明细表导入excel文件
        wb_result = load_workbook(filename=wkfj_result_import_path)  # 打开结果表导入excel文件
        # 把流程中的样本号写入待导入模板
        for i in range(1, len(samples) + 1):
            k = i + 1
            self.sleep(0.5)
            samples_lims_id = self.get_text('css', all_sample_lims.format(i))
            samples_lab_id = self.get_text('css', all_sample_lab.format(i))
            ws = wb.active
            ws2 = wb_result.active
            # 根据需要修改表格数据
            fjwkmc = 'NJ-' + str_time + '-' + probe + 'FJ01'
            ws.cell(k, 1, samples_lims_id)  # 修改第k行，第index列值,lims号
            ws.cell(k, 2, samples_lab_id)  # 修改第k行，第index列值，实验室号
            ws.cell(k, 3, fjwkmc)  # pooling号
            ws2.cell(k, 1, fjwkmc)  # 把富集名称同时写入结果表

            # 为每个样本生成不同富集名称

            # if j < 10:
            #     fjwkmc = 'NJ-' + str_time + '-' + probe + 'FJ0{}'  # 富集名称
            #     ws.cell(k, 3, fjwkmc.format(j))  # pooling号
            #     ws2.cell(k, 1, fjwkmc.format(j))  # 把富集名称同时写入结果表
            # elif j >= 0:
            #     fjwkmc = 'NJ-' + str_time + '-' + probe + 'FJ{}'  # 富集名称
            #     ws.cell(k, 3, fjwkmc.format(j))  # pooling号
            #     ws2.cell(k, 1, fjwkmc.format(j))  # 把富集名称同时写入结果表
            # elif j < 20:
            #     fjwkmc = 'NJ-' + str_time + '-101-B-1@{}'  # 富集名称
            #     ws.cell(k, 3, fjwkmc.format(int(j) - 10))  # pooling号
            #     ws2.cell(k, 1, fjwkmc.format(int(j) - 10))  # 把富集名称同时写入结果表
            # elif j < 30:
            #     fjwkmc = 'NJ-' + str_time + '-101-C-1@{}'  # 富集名称
            #     ws.cell(k, 3, fjwkmc.format(int(j) - 20))  # pooling号
            #     ws2.cell(k, 1, fjwkmc.format(int(j) - 20))  # 把富集名称同时写入结果表
            # elif j < 40:
            #     fjwkmc = 'NJ-' + str_time + '-101-D-1@{}'  # 富集名称
            #     ws.cell(k, 3, fjwkmc.format(int(j) - 30))  # pooling号
            #     ws2.cell(k, 1, fjwkmc.format(int(j) - 30))  # 把富集名称同时写入结果表
            # elif j < 50:
            #     fjwkmc = 'NJ-' + str_time + '-101-E-1@{}'  # 富集名称
            #     ws.cell(k, 3, fjwkmc.format(int(j) - 40))  # pooling号
            #     ws2.cell(k, 1, fjwkmc.format(int(j) - 40))  # 把富集名称同时写入结果表
            # elif j < 60:
            #     fjwkmc = 'NJ-' + str_time + '-101-F-1@{}'  # 富集名称
            #     ws.cell(k, 3, fjwkmc.format(int(j) - 50))  # pooling号
            #     ws2.cell(k, 1, fjwkmc.format(int(j) - 50))  # 把富集名称同时写入结果表
        wb.save(wkfj_detail_import_path)
        wb_result.save(wkfj_result_import_path)
        self.sleep(1)
        # 从模板中复制出修改后的待导入数据
        data = xlrd.open_workbook(wkfj_detail_import_path)
        num_list = []
        for index in range(0, len(samples)):
            h = index + 1
            tables = data.sheets()[0]
            vals = tables.row_values(h)
            imp_data = '\t'.join(map(str, vals))
            num_list.append(imp_data)
        print("\n".join(map(str, num_list)))
        pyperclip.copy("\n".join(map(str, num_list)))

    # 返回明细表
    def goback_detail(self):
        """返回明细表"""
        urldata = read_yaml(functionpageURL_path)
        log.info('点击按钮返回明细表')
        self.click_by_js('css', goback_detail)
        self.clicks('css', goback_page_info)
        self.wait_loading()

        url = self.get_current_url()  # 获取当前页面URL地址
        print('获取的URL地址', url)
        urldata["url"] = url
        with open(functionpageURL_path, 'w', encoding='utf-8') as fs:  # 写入模式获取的URL地址到yaml文件中
            yaml.safe_dump(urldata, fs, allow_unicode=True)
        print("写入后的URL地址", urldata)

    # 明细表入库信息、批量粘贴导入实验数据
    def detail_libraryenrichment(self, probe):
        """文库富集明细表，选择入库信息、批量包装余量、录入96孔版位置,无探针任务类型"""
        self.clicks('css', generatedSortNo)  # 自动生成序号
        self.sleep(0.5)
        self.click_by_js('css', detail_all_choice)  # 列表全选按钮
        log.info("文库富集明细表批量粘贴导入实验数据")
        self.read_import_data(probe)
        self.clicks('css', batch_paste_import_package)  # 明细表批量粘贴导入按钮
        self.sleep(0.5)
        self.clicks('css', batch_paste_import_package_textarea)
        self.findelement('css', batch_paste_import_package_textarea).send_keys(Keys.CONTROL, 'v')
        self.sleep(0.5)
        self.clicks('css', batch_paste_import_package_comfirm)  # 明细表批量粘贴导入弹框确认按钮
        self.sleep(0.5)
        log.info("富集明细表自动计算")
        self.clicks('xpath', automatic)  # 自动计算
        self.sleep(0.5)
        # if self.isClickable('css',tips):
        #     self.clicks('css', tips)
        log.info("富集明细表保存页面数据操作")
        self.clicks('css', detail_save_result)
        self.wait_loading()

    # 明细表入库信息、批量粘贴导入实验数据
    def detail_libraryenrichment_probe(self, probe):
        """文库富集明细表，批量粘贴导入实验数据,有探针任务类型"""
        self.click_by_js('css', detail_all_choice)  # 列表全选按钮
        self.sleep(0.5)
        self.clicks('css', generatedSortNo)  # 自动生成序号

        log.info("文库富集明细表批量粘贴导入实验数据")
        self.read_import_data(probe)
        self.clicks('css', batch_paste_import_package)  # 明细表批量粘贴导入按钮
        self.sleep(0.5)
        self.clicks('css', batch_paste_import_package_textarea)
        self.findelement('css', batch_paste_import_package_textarea).send_keys(Keys.CONTROL, 'v')
        self.sleep(0.5)
        self.clicks('css', batch_paste_import_package_comfirm)  # 明细表批量粘贴导入弹框确认按钮
        self.sleep(0.5)
        # 在表单中选择探针和录入进入量（实际）的值
        all_sample = self.findelements('css', all_samples)
        for i in range(1, len(all_sample) + 1):
            self.clicks('css', probe_select.format(i))
            self.clicks('css', probe_select_input.format(i))
            self.sleep(0.5)
            self.clicks('css', probe_choice)
            self.sleep(0.5)
            self.clicks('css', theoreticalUsedAmt.format(i))
            self.input('css', theoreticalUsedAmt_input.format(i), 5)
            self.sleep(0.5)

        log.info("富集明细表自动计算")
        self.clicks('xpath', automatic)  # 自动计算
        self.sleep(0.5)
        if self.isDisplayed('css', tips):
            self.clicks('css', tips)
        self.sleep(0.5)
        log.info("富集明细表保存页面数据操作")
        self.clicks('css', detail_save_result)
        self.wait_loading()

    # 明细表生成结果，保存数据
    def detail_create_result(self, index):
        log.info("批量数据后生成结果")
        self.click_by_js('css', detail_all_choice)  # 列表全选按钮
        self.sleep(1)
        self.click_by_js('css', create_result)
        self.wait_loading()
        Screenshot(self.driver).get_img("文库富集明细表生成结果", "生成结果成功")
        self.executeJscript('document.getElementsByClassName("vxe-table--body-wrapper")[0].scrollLeft=6000')
        self.sleep(1)
        result_info = self.get_text('css', result_status.format(index))
        return result_info

    # 明细表提交
    def detail_submit(self):
        """文库富集明细表，样本提交操作"""
        log.info("文库富集明细表样本提交")
        self.click_by_js('css', detail_all_choice)  # 列表全选按钮
        self.sleep(0.5)
        self.clicks('css', submit_btn)  # 提交按钮
        self.sleep(0.5)
        Screenshot(self.driver).get_img("文库富集明细表点击提交按钮", "弹出提交确认按钮")
        self.clicks('css', submit_comfirm)  # 提交弹框确认按钮
        self.wait_loading()

    # 明细表入库
    def detail_into_storage(self, nub):
        """文库富集明细表样本入库操作"""
        log.info('文库富集明细表，样本入库操作')
        self.clicks('css', detail_all_choice)  # 列表全选按钮
        self.sleep(0.5)
        self.clicks('css', deposit_into_storage)  # 入库按钮
        self.sleep(0.5)
        self.clicks('css', storage_all_choice)  # 入库弹框全选按钮
        self.sleep(0.5)
        log.info('文库富集明细表，样本入库选择入样本盒')
        self.clicks('css', batch_paste_sample_box)  # 入库弹框选择样本盒按钮
        self.wait_loading()
        self.input('css', target_storage, '自动化测试用(勿删)')
        self.sleep(0.5)
        self.clicks('css', select_sample_box_search)
        self.wait_loading()
        self.clicks('css', select_sample_box_choice)  # 入库弹框选选择样本盒值，默认选择列表第一条数据
        self.clicks('css', select_sample_box_comfirm)  # 入库弹框选选择样本盒弹框，确认按钮
        self.sleep(1)

        log.info('文库富集明细表，样本入库录入盒内位置')
        taskstatus = self.get_text('css', detail_task_id)  # 获取任务单号
        lims_id = executeSql.test_select_limsdb(wkfj_detail_sql1.format(taskstatus[5:].strip()))  # 从数据库获取当前任务单号下样本lims号
        lims_list = [item[key] for item in lims_id for key in item]  # 把获取的lims号转换为一维列表
        nub_list = [str(i) for i in range(1, len(lims_list) + 1)]  # 根据lims样本数量，生成数字列表，作为盒内位置编号用
        res = [list(i) for i in zip(lims_list, nub_list)]  # 将lims号和数字编号转换为二维列表格式，写入Excel
        print(res)
        pandas_write_excel(res, position_in_box_path)  # 把样本号和盒内位置编号写入Excel模板

        data = xlrd.open_workbook(position_in_box_path)  # 从Excel读取模板样本号和盒内位置编号
        num_list = []
        for index in range(0, len(lims_list)):
            tables = data.sheets()[0]
            # allrows = tables.nrows
            vals = tables.row_values(index)
            imp_data = '\t'.join(map(str, vals))
            num_list.append(imp_data)
        print("\n".join(map(str, num_list)))
        pyperclip.copy("\n".join(map(str, num_list)))

        # 粘贴到【批量粘贴盒内位置】文本框
        self.clicks('css', batch_copy_BoxPosition)
        self.findelement('css', batch_copy_BoxPosition_input).send_keys(Keys.CONTROL, 'v')
        self.sleep(0.5)
        self.clicks('css', batch_copy_BoxPosition_comfirm)
        self.sleep(1)
        Screenshot(self.driver).get_img("文库富集明细表点击入库按钮，在弹框中录入库位信息和盒内位置后点击下一步", "样本入库成功")

        self.clicks('css', storage_next)
        self.wait_loading()
        self.executeJscript('document.getElementsByClassName("vxe-table--body-wrapper")[0].scrollLeft=6000')
        self.sleep(1)
        pageinfo = self.get_text('css', detail_sumbit_status.format(nub))  # 入库后获取样本状态，判断是否入库成功
        print(pageinfo)
        return pageinfo

    # 结果表批量数据录入
    def ultrasonic_result_data_input(self):
        """文库富集结果表修改产物类型、批量数据"""
        samples = self.findelements('css', result_samples_for_total)
        # 从模板中复制出修改后的待导入数据
        data = xlrd.open_workbook(wkfj_result_import_path)
        num_list = []
        for index in range(0, len(samples)):
            h = index + 1
            tables = data.sheets()[0]
            vals = tables.row_values(h)
            imp_data = '\t'.join(map(str, vals))
            num_list.append(imp_data)
        pyperclip.copy("\n".join(map(str, num_list)))

        self.clicks('css', result_batch_paste_import)
        self.sleep(0.5)
        self.clicks('css', result_batch_paste_import_package_textarea)
        self.findelement('css', result_batch_paste_import_package_textarea).send_keys(Keys.CONTROL, 'v')
        self.sleep(0.5)
        self.clicks('css', result_batch_paste_import_package_comfirm)  # 结果表批量粘贴导入弹框确认按钮
        self.sleep(0.5)

    # 结果表提交
    def result_submit_sample(self):
        """结果表提交,无探针富集"""
        log.info(' 文库富集结果表提交操作')
        self.clicks('css', result_all_choice)
        self.sleep(0.5)
        self.clicks('css', result_submit)  # 提交按钮
        self.wait_loading()

        Screenshot(self.driver).get_img("无探针文库富集结果表点击提交按钮", "弹出提交确认按钮")

        self.clicks('css', result_submit_comfirm)  # 提交确认按钮
        self.wait_loading()
        # 页面拉到【是否已提交】表单，获取表单文本，进行断言
        pageinfo = self.get_text('css', result_sumbit_status)
        print(pageinfo)
        return pageinfo

    # 结果表提交
    def result_submit_sample_probe(self):
        """结果表提交,有探针富集"""
        log.info('有探针类型文库富集结果表提交操作')
        all_sample = self.findelements('css', result_samples_for_total)
        for i in range(1, len(all_sample) + 1):
            self.clicks('css', consistenceAmtOne.format(i))
            self.input('css', consistenceAmtOne_input.format(i), 2)
            self.sleep(0.5)
            self.clicks('css', consistenceAmtTwo.format(i))
            self.input('css', consistenceAmtTwo_input.format(i), 2)
            self.sleep(0.5)

        self.clicks('xpath', result_automatic)
        self.sleep(0.5)
        self.clicks('css', result_submit)  # 提交按钮
        self.wait_loading()

        Screenshot(self.driver).get_img("有探针文库富集结果表点击提交按钮", "弹出提交确认按钮")

        self.clicks('css', result_submit_comfirm)  # 提交确认按钮
        self.wait_loading()
        # 页面拉到【是否已提交】表单，获取表单文本，进行断言
        pageinfo = self.get_text('css', result_sumbit_status)
        print(pageinfo)
        return pageinfo

    # 结果表样本流程环节写入Excel
    def write_data_to_excel(self):
        """根据结果表样本下一步流程，把对应的样本lims号、实验室号、下一步流程以追加形式写入该流程的Excel"""
        self.add_excel_xlxs(fj_next_step, pooling_result, result_task_id)
        print("下一步流程写入成功")

    # 完成任务单
    def complete_task(self):
        """完成任务单"""
        log.info(' 文库富集结果表,点击完成任务单')
        self.clicks('css', enter_result_list_btn)
        self.wait_loading()

        self.clicks('css', result_complete_task_btn)
        self.wait_loading()
        # 调用自定义截图方法
        Screenshot(self.driver).get_img("文库富集结果表点击完成任务单按钮", "完成任务单成功，状态改为完成")

        taskstatus = self.get_text('css', task_status)
        print(taskstatus)
        return taskstatus

    # 文库富集任务列表查询样本所在任务单
    def serach_task(self, path):
        """首页面查询已完成的样本任务单"""
        lims_id = read_excel_col(path, 'lims号')
        self.clicks('css', search)
        self.sleep(1)
        self.input('css', search_lims_sample_num, lims_id[0])
        self.sleep(1)
        self.clicks('css', search_confirm)
        self.wait_loading()
        samples = self.findelements('xpath', sample_page_list)
        print(len(samples))
        return len(samples)
