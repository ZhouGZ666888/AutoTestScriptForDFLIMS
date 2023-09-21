# -*- coding: utf-8 -*-
# @Time    : 2022/09/14
# @Author  : guanzhong.zhou
# @File    : QPCR页面功能封装
import pyperclip, xlrd, os
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
from common.editYaml import read_yaml, save_yaml
from common.DataBaseConnection import executeSql
from common.xlsx_excel import pandas_write_excel, read_excel_col, get_lims_for_excel_by_rows, get_lims_for_excel_by_col
from common.screenshot import Screenshot
from PageElements.qpcr_ele import *
from conf.all_path import qpcr_dxk_file_path, wkgj_file_path, functionpageURL_path, \
    position_in_box_path, excel_doc_file_path, mNGTaskSampleId
from conf.execute_sql_action import qpcr_sql
from uitestframework.basepageTools import BasePage
from common.logs import log
from uitestframework.exceptionsTools import ElementNotFound


class QpcrPage(BasePage):
    """QPCR页面基础方法"""

    def add_qpcr_task(self):
        """新增qpcr任务"""
        log.info('新增qpcr任务')
        self.clicks('css', add_sample_QPCR_task)
        self.wait_loading()

    def to_be_select(self, taskType):
        """待选表功能封装"""
        log.info('获取qpcr任务lims号')

        try:
            log.info('选择任务类型：{}'.format(taskType))
            self.clicks('css', task_type)
            self.sleep(0.5)
            print(task_type_choice.format(taskType))
            self.clicks('xpath', task_type_choice.format(taskType))
            self.sleep(0.5)

            log.info('选择sop')
            self.click_by_js('css', task_sop)
            self.clicks('css', task_sop_choice)
            self.sleep(1)

            log.info('核对lims样本号')
            self.clicks('xpath', check_lims_sample_num)
            if taskType == 'mNGS':
                seqSampleId = read_yaml(mNGTaskSampleId)['mNGSinspectTaskSampleId']  # mNGS样本号
                print(seqSampleId)
                self.input('css', check_lims_sample_number_textarea, seqSampleId)
            elif taskType == '迪迅康':
                limsnub = get_lims_for_excel_by_col(qpcr_dxk_file_path)  # 迪讯康样本号
                print(limsnub)
                self.input('css', check_lims_sample_number_textarea, limsnub)
            elif taskType == '其他':
                otherNub = get_lims_for_excel_by_rows(wkgj_file_path, 0, -5, 'lims号')  # 其他类型样本号
                print(otherNub)
                self.input('css', check_lims_sample_number_textarea, otherNub)
            self.clicks('xpath', check_lims_sample_number_confirm)
            self.sleep(0.5)
        except ElementNotFound:
            log.error('元素定位失败')

    def add_sample_to_detail(self):
        """选中样本进入明细表"""
        self.clicks('css', all_choice)
        self.sleep(0.5)

        log.info('选择样本加入明细表')
        self.clicks('css', addSelect_or_save_btn)
        self.wait_loading()

        Screenshot(self.driver).get_img("文库构建待选表核对lims号功能，并保存任务单号")
        pageinfo = self.get_pageinfo()
        self.wait_loading()

    def enter_function_page(self, url):
        """进入指定url页面"""
        js = 'window.open("{}");'
        self.executeJscript(js.format(url))
        self.wait_loading()

    # 进入明细表或结果表
    def enter_result_list(self, ele, page):
        """
        待选表或者明细表，在进入下一节点时，获取下一节点URL地址，并写入临时数据文件中
        :param ele: 点击进入下一节点元素定位
        :param page: 下一节点页面名称
        """
        urldata = read_yaml(functionpageURL_path)

        self.clicks('css', ele)
        log.info('点击按钮进入{}'.format(page))
        self.wait_loading()

        url = self.get_current_url()  # 获取当前页面URL地址
        print('获取的URL地址', url)
        urldata["url"] = url
        save_yaml(functionpageURL_path, urldata)  # 写入模式获取的URL地址到yaml文件中
        print("写入后的URL地址", urldata)

    def qpcr_detail(self):
        """QPCR明细表操作"""
        log.info('明细表选中样本，点击批量计算')
        self.clicks('css', detail_all_choice)
        self.sleep(0.5)
        self.clicks('css', auto_complete)
        self.sleep(0.5)

    def detail_sumbit(self):
        """明细表提交"""
        log.info('明细表提交')
        self.clicks('xpath', sumbit_btn)
        self.wait_loading()
        self.clicks('css', submit_comfirm)
        self.wait_loading()

    def qpcr_detail_addncpc(self):
        """qpcr明细表添加NC/PC"""
        log.info('QPCR明细表，添加NC/PC')
        self.clicks('xpath', add_nc_pc)
        self.wait_loading()

    def detail_into_storage(self):
        """明细表样本入库操作"""
        try:
            log.info('明细表，样本入库操作')
            self.click_by_js('css', detail_all_choice)
            self.sleep(1)
            self.clicks('xpath', deposit_into_storage)  # 入库按钮
            self.sleep(1)
            self.clicks('css', storage_all_choice)  # 入库弹框全选按钮
            self.sleep(1)

            log.info('QPCR明细表，样本入库选择入样本盒')
            self.clicks('css', batch_paste_sample_box)  # 入库弹框选择样本盒按钮
            self.wait_loading()
            self.input('css', target_storage, '自动化测试用(勿删)')
            self.sleep(0.5)
            self.clicks('css', select_sample_box_search)
            self.wait_loading()
            self.clicks('css', select_sample_box_choice)  # 入库弹框选选择样本盒值，默认选择列表第一条数据
            self.sleep(0.5)
            self.clicks('xpath', select_sample_box_comfirm)  # 入库弹框选选择样本盒弹框，确认按钮
            self.sleep(0.5)

            log.info('QPCR明细表，样本入库批量粘贴盒内位置')
            taskstatus = self.get_text('css', detail_task_id)  # 获取任务单号
            lims_id = executeSql.test_select_limsdb(
                qpcr_sql.format(taskstatus.strip()[-18:]))  # 从数据库获取当前任务单号下样本lims号

            lims_list = [item[key] for item in lims_id for key in item]  # 把获取的lims号转换为一维列表
            nub_list = [str(i) for i in range(1, len(lims_list) + 1)]  # 根据lims样本数量，生成数字列表，作为盒内位置编号用
            res = [list(i) for i in zip(lims_list, nub_list)]  # 将lims号和数字编号转换为二维列表格式，写入Excel
            pandas_write_excel(res, position_in_box_path)

            data = xlrd.open_workbook(position_in_box_path)
            num_list = []
            for B in range(0, len(lims_list)):
                tables = data.sheets()[0]
                vals = tables.row_values(B)
                imp_data = '\t'.join(map(str, vals))
                num_list.append(imp_data)
            print("\n".join(map(str, num_list)))
            pyperclip.copy("\n".join(map(str, num_list)))

            # 粘贴到【批量粘贴盒内位置】文本框
            self.click_by_js('css', batch_copy_BoxPosition)
            self.sleep(0.5)
            self.findelement('xpath', batch_copy_BoxPosition_input).send_keys(Keys.CONTROL, 'v')
            self.sleep(1)
            self.clicks('xpath', batch_copy_BoxPosition_comfirm)
            self.sleep(0.5)

            # 调用自定义截图方法
            Screenshot(self.driver).get_img("QPCR明细表入库")

            self.clicks('xpath', storage_next)
            self.wait_loading()

            self.executeJscript('document.getElementsByClassName("vxe-table--body-wrapper")[0].scrollLeft=3080')
            samples_status = self.get_text('css', submit_status)
            print('明细表状态', samples_status)
            self.sleep(0.5)
            return samples_status
        except Exception as info:
            print(info)

    def qpcr_result_export_out_btn(self):
        """QPCR结果表，导出复孔信息"""
        log.info('QPCR结果表，导出复孔信息')
        self.clicks('xpath', export_out_btn)
        self.wait_loading()

    def qpcr_result_export_in_btn(self):
        """QPCR结果表，导入复孔信息"""
        log.info('QPCR结果表，导入复孔信息')

        log.info('在设置的下载文件夹下，获取最新导出的复孔信息文件')
        lists = os.listdir(excel_doc_file_path)
        lists.sort(key=lambda fn: os.path.getmtime(excel_doc_file_path + '\\' + fn))
        filepath = os.path.join(excel_doc_file_path, lists[-1])

        wb = load_workbook(filename=filepath)  # 打开excel文件
        # 获取总打开Excel文件总行数
        table1 = wb.worksheets[0]
        nrows = table1.max_row
        print('复孔信息文件行数', nrows)
        # 把Cq值、Melt Peck值（实测）写入复孔文件
        for i in range(1, nrows):
            k = i + 1
            ws = wb.active
            # 根据需要修改表格数据
            ws.cell(k, 6, 3)  # 修改第k行，第index列值
            ws.cell(k, 7, 85.5)
        wb.save(filepath)
        self.sleep(1)

        # 从模板中复制出修改后的待导入数据
        data = xlrd.open_workbook(filepath)
        num_list = []
        for index in range(0, nrows - 1):
            h = index + 1  # xlrd行下标从0开始
            tables = data.sheets()[0]
            vals = tables.row_values(h)
            imp_data = '\t'.join(map(str, vals))
            num_list.append(imp_data)
        print("\n".join(map(str, num_list)))
        pyperclip.copy("\n".join(map(str, num_list)))

        self.clicks('xpath', export_in_btn)
        self.wait_loading()
        self.clicks('css', export_in_input)
        self.findelement('css', export_in_input).send_keys(Keys.CONTROL, 'v')
        self.sleep(1)
        self.clicks('css', export_in_input_confirm)  # 明细表批量粘贴导入弹框确认按钮
        self.sleep(1)
        self.clicks('css', export_in_input_confirm)
        self.wait_loading()

    def result_auto_Complete(self):
        """结果表自动判读"""
        log.info('QPCR结果表，自动判读')
        self.clicks('css', result_all_choice)
        self.sleep(0.5)
        self.clicks('xpath', sumbit_auto_Complete)
        self.wait_loading()

    def result_submit(self, index):
        """结果表提交"""
        log.info('QPCR结果表，提交')
        self.clicks('css', result_all_choice)
        self.sleep(0.5)
        self.clicks('xpath', result_sumbit_btn)
        self.wait_loading()
        samples_status = self.get_text('css', sample_status.format(index))
        print('结果表状态', samples_status)
        self.sleep(1)
        return samples_status

    def go_back_detail(self):
        """返回明细表，完成PC/NC任务提交"""
        self.sleep(1)
        self.click_by_js('xpath', go_back)
        self.sleep(0.5)
        self.clicks('css', go_back_confirm)
        self.wait_loading()

        self.qpcr_detail()  # 自动计算
        self.detail_sumbit()  # 提交
        self.wait_loading()
        self.click_by_js('css', go_result)  # 进入结果表
        self.wait_loading()
        self.sleep(1)

    def complete_task(self):
        """完成任务单"""
        log.info(' qpcr结果表,点击完成任务单')

        self.click_by_js('xpath', result_complete_task_btn)
        self.wait_loading()
        task_status = self.get_text('css', result_task_status)
        return task_status

    def serach_task(self, path):
        """首页面查询已完成的样本任务单"""
        lims_id = read_excel_col(path, 'lims号')

        self.sleep(1)
        self.input('css', search_lims_sample_num, lims_id[0])
        self.sleep(1)
        self.clicks('css', search_btn)

        self.wait_loading()
        samples = self.findelements('css', sample_page_list)
        return len(samples)

    # 获取页面提示信息
    def get_pageinfo(self):
        """
        获取页面操作提示信息
        Task list saved successfully---保存样本到任务单成功
        Submit successfully---提交成功
        sample in storage successfully---入库成功
        """
        if self.isElementExists('xpath', page_success_info):
            return self.get_text('xpath', page_success_info)
        else:
            return None
